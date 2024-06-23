from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from .models import Event
from .forms import EventForm
from django.core.paginator import Paginator
import xlsxwriter
from django import forms
from django.http import HttpResponse
import os
import tempfile
import pandas as pd
import openpyxl
from openpyxl import Workbook
from fpdf import FPDF



def events(request):
    events = Event.objects.filter(association=request.user.member.logged_in_association)
    paginator = Paginator(events, 10)
    page = request.GET.get('page')
    events = paginator.get_page(page)
    context = {'events': events}
    return render(request, 'benevofy/events.html', context)



def delete_event(request, event_id):
    event = Event.objects.get(id=event_id)
    event.delete()
    return redirect('events')


def update_event(request, event_id):
    event = Event.objects.get(id=event_id)
    if request.method == 'POST':
        form = EventForm(request.POST, instance=event)
        if form.is_valid():
            form.save()
            return render(request, 'benevofy/events.html')
    else:
        form = EventForm(instance=event)
    return render(request, 'benevofy/update_event.html', {'form': form})


def create_event(request):
    if request.method == 'POST':
        form = EventForm(request.POST)
        if form.is_valid():
            form.instance.association = request.user.member.logged_in_association
            event = form.save()
            event.association.events.add(event)
            return redirect('events')
    else:
        association = request.user.member.logged_in_association
        form = EventForm(initial={'association': association})
        form.fields['association'].widget = forms.HiddenInput()
    return render(request, 'benevofy/create_event.html', {'form': form})


def event_report(request, event_id):
    event = Event.objects.get(id=event_id)
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    format = request.GET.get('format')

    contributions = event.contributions.filter(created_at__gte=from_date, created_at__lte=to_date)

    # Create a DataFrame and export it to the specified format
    df = pd.DataFrame({
        'Member': [contribution.user.user.get_full_name() for contribution in contributions],
        'Amount': [contribution.amount for contribution in contributions],
        'Reference': [contribution.reference for contribution in contributions],
        'Date': [contribution.created_at.strftime('%Y-%m-%d') for contribution in contributions]
    })

    if format == 'xlsx':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=event_report_{event.name}_{from_date}_{to_date}.xlsx'
        df.to_excel(response, index=False)
    elif format == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        filename = f'event_report_{event.name}_{from_date}_{to_date}.pdf'
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
        pdf = FPDF(format='A4')
        pdf.add_page()
        pdf.set_xy(10, 10)
        pdf.set_font('Helvetica', '', 16)
        pdf.cell(280, 20, 'Event Report', 0, 1, 'C')
        pdf.cell(280, 20, f'{event.name} - {from_date} to {to_date}', 0, 1, 'C')
        pdf.cell(280, 20, '', 0, 1)
        pdf.set_font('Helvetica', '', 10)
        pdf.cell(60, 10, 'Member', 1, 0)
        pdf.cell(60, 10, 'Amount', 1, 0, 'C')
        pdf.cell(60, 10, 'Reference', 1, 0)
        pdf.cell(60, 10, 'Date', 1, 1, 'C')
        for _, row in df.iterrows():
            pdf.cell(60, 10, row['Member'], 1, 0)
            pdf.cell(60, 10, f'{row["Amount"]:0.2f}', 1, 0, 'R')
            pdf.cell(60, 10, row['Reference'], 1, 0)
            pdf.cell(60, 10, row['Date'], 1, 1, 'C')
        pdf_content = pdf.output().encode('latin-1')
        response.write(pdf_content)
    else:
        raise ValueError('Invalid format')

    return response



def close_event(request, event_id):
    event = Event.objects.get(id=event_id)
    event.status = 'Closed'
    event.save()
    return redirect('events')


def resume_event(request, event_id):
    event = Event.objects.get(id=event_id)
    event.status = 'Active'
    event.save()
    return redirect('events')
